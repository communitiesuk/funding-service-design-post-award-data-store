"""
This script processes files in a directory and performs actions based on the specified command-line arguments.
The processed files are either posted to an endpoint or stored in separate directories, depending on the chosen option.

Usage:
    python script.py [directory_path] [endpoint_url] [-d | -e]

Arguments:
    directory_path (str): Path to the directory containing the files to be processed.
    endpoint_url (str): URL of the endpoint to post the files to.

Options:
    -d, --directories: Store the results in separate directories.
    -e, --excel: Store the results in an Excel file.

Examples:
    python script.py /path/to/files/ http://example.com/api/ -d
    python script.py /path/to/files/ http://example.com/api/ -e

Dependencies:
    - pandas
    - requests
    - openpyxl

Note:
    - The script assumes that the files to be processed are located in the specified directory.
    - The output will be saved in a subdirectory named 'output' within the specified directory.
"""

import argparse
import json
import os
from datetime import datetime

import pandas as pd
import requests
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


def post_file(file_path, endpoint_url):
    """
    Post a file to the specified endpoint URL.

    :param file_path: Path to the file to be posted.
    :param endpoint_url: URL of the endpoint to post the file to.
    :return: Response from the server.
    """
    with open(file_path, "rb") as file:
        response = requests.post(
            endpoint_url,
            files={
                "excel_file": (file.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            },
            data={"reporting_round": 3},
        )
    return response


def save_response(file_path, response, output_folder):
    """
    Save the response from the server to a file in the output folder.

    :param file_path: Path to the file for which the response was received.
    :param response: Response from the server.
    :param output_folder: Path to the output folder.
    """
    file_name = os.path.basename(file_path)
    folder_name = os.path.splitext(file_name)[0]
    folder_path = os.path.join(output_folder, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    response_file_path = os.path.join(folder_path, "response.txt")
    with open(response_file_path, "w") as file:
        file.write(response.text)


def output_to_excel():
    """
    Process files in a directory and store the results in an Excel file.
    """
    output_df = pd.DataFrame(columns=["File Name", "Success", "Errors"])

    dir_items = [item for item in os.listdir(args.directory_path) if item.endswith(".xlsx")]
    for idx, file_name in enumerate(dir_items):
        if file_name == "output":
            continue

        print(
            f"Testing file {idx}/{len(dir_items) - 1} - {file_name}. \nSuccess ({output_df['Success'].sum()}), "
            f"Failed ({len(output_df) - output_df['Success'].sum()})"
        )

        file_path = os.path.join(args.directory_path, file_name)
        if os.path.isfile(file_path):
            response = post_file(file_path, args.endpoint_url)

            status_code = response.status_code
            success = status_code == requests.codes.ok
            if success:
                errors = ""
            elif status_code == 400:  # validation error
                errors = json.dumps(response.json()["validation_errors"])
            else:  # internal error e.g. DB error
                errors = json.dumps(response.json())

            output_df.loc[len(output_df)] = {"File Name": file_name, "Success": success, "Errors": errors}
            print("Success." if success else "Failed.")
        print()

    workbook = Workbook()
    sheet = workbook.active
    sheet.column_dimensions["C"].width = 160
    for row in dataframe_to_rows(output_df, index=False, header=True):
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        row[2].alignment = Alignment(wrap_text=True)
    workbook.save(output_folder + f"/results_{datetime.now().strftime('%Y%m%dT%H%M%S%z')}.xlsx")


def output_to_directories():
    """
    Process files in a directory and store the results in a set of folders named after each file, along with 2 text
    files with names of successfully and unsuccesfully processed files.
    """
    success_responses = []
    failure_responses = []

    dir_items = os.listdir(args.directory_path)
    for idx, file_name in enumerate(dir_items):
        if file_name == "output":
            continue

        print(
            f"Testing file {idx}/{len(dir_items)} - {file_name}. \nSuccess ({len(success_responses)}), "
            f"Failed ({len(failure_responses)})"
        )

        file_path = os.path.join(args.directory_path, file_name)
        if os.path.isfile(file_path):
            response = post_file(file_path, args.endpoint_url)

            if response.status_code == requests.codes.ok:
                success_responses.append(file_name)
                print("Success.")
            else:
                failure_responses.append(file_name)
                save_response(file_path, response, output_folder)
                print("Failed.")
        print()

    with open(os.path.join(output_folder, "success_responses.txt"), "w") as success_file:
        success_file.write("\n\n".join(success_responses))

    with open(os.path.join(output_folder, "failure_responses.txt"), "w") as failure_file:
        failure_file.write("\n\n".join(failure_responses))


parser = argparse.ArgumentParser(description="Process files and post to an endpoint.")
parser.add_argument("directory_path", type=str, help="Path to the directory containing the files.")
parser.add_argument("endpoint_url", type=str, help="URL of the endpoint to post the files to.")

group = parser.add_mutually_exclusive_group(required=True)
group.add_argument("-d", "--directories", action="store_true", help="Store results to separate directories")
group.add_argument("-e", "--excel", action="store_true", help="Store results to excel")

args = parser.parse_args()

output_folder = os.path.join(args.directory_path, "output")
os.makedirs(output_folder, exist_ok=True)

if args.directories:
    output_to_directories()
elif args.excel:
    output_to_excel()
